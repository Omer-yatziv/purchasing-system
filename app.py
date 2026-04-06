import streamlit as st
import pandas as pd
from thefuzz import fuzz, process
from io import BytesIO
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="מערכת רכש חכמה", layout="wide")
st.title("🎯 מערכת הצלבת רכש - גרסת פורמט מקורי")

@st.cache_data
def load_erp():
    for ext in ['xlsx', 'csv']:
        try:
            name = f"erp_master.{ext}"
            df = pd.read_excel(name) if ext == 'xlsx' else pd.read_csv(name)
            df.columns = df.columns.str.strip()
            return df
        except: continue
    return None

erp_df = load_erp()

if erp_df is None:
    st.error("⚠️ קובץ ה-ERP חסר ב-GitHub")
else:
    st.success(f"ERP נטען בהצלחה")
    purchase_file = st.file_uploader("העלה בקשת רכש (המקורית)", type=['xlsx', 'xlsm'])
    
    if purchase_file:
        skip_rows = st.number_input("באיזו שורה מתחילה הטבלה (כותרות)?", value=10) - 1
        
        # טעינת הקובץ המקורי עם openpyxl כדי לשמור על עיצוב
        wb = openpyxl.load_workbook(purchase_file)
        ws = wb.active # הלשונית הראשונה
        
        # קריאת הנתונים לעיבוד (מתוך ה-Worksheet)
        data = ws.values
        cols = next(data)[0:] # קריאת הכותרות
        df_full = pd.DataFrame(data, columns=cols)
        # סינכרון עם שורת הכותרת האמיתית
        df_for_logic = pd.read_excel(purchase_file, header=skip_rows)
        df_for_logic.columns = df_for_logic.columns.astype(str).str.strip()

        st.info("בדיקת עמודות מקור: B-מס\"ד, C-קבוצה, D-ח\"ג")
        
        if st.button("הפעל עיבוד ושמור פורמט"):
            results_for_analysis = []
            erp_choices = erp_df['תיאור פריט'].astype(str).tolist()
            
            # הכנת עמודות חדשות בגיליון המקורי (הזזת הכל ימינה כדי לפנות מקום ל-3 עמודות ERP)
            ws.insert_cols(1, amount=3)
            ws.cell(row=skip_rows+1, column=1).value = "קוד פריט סאפ"
            ws.cell(row=skip_rows+1, column=2).value = "תיאור סאפ"
            ws.cell(row=skip_rows+1, column=3).value = "% התאמה"

            # עיבוד שורה שורה (החל משורה שאחרי הכותרת)
            current_row_excel = skip_rows + 2
            
            for idx, row in df_for_logic.iterrows():
                # בניית שאילתה (מבטיח שימוש בעמודות שציינת)
                q_group = str(row.get('קבוצה', ''))
                q_mat = str(row.get('ח"ג', ''))
                q_desc = str(row.get('תאור/מידה', ''))
                
                if q_desc == "nan" or q_desc == "": 
                    current_row_excel += 1
                    continue
                
                query = f"{q_group} {q_mat} {q_desc}".replace("nan", "").strip()
                
                # מציאת 4 התאמות הכי טובות (הראשונה היא הנבחרת, היתר הצעות חלופיות)
                matches = process.extract(query, erp_choices, scorer=fuzz.token_set_ratio, limit=4)
                
                if matches and matches[0][1] > 40:
                    best_match_text, score = matches[0][0], matches[0][1]
                    erp_row = erp_df[erp_df['תיאור פריט'] == best_match_text].iloc[0]
                    
                    # כתיבה לגיליון האקסל המקורי בעמודות החדשות A, B, C
                    ws.cell(row=current_row_excel, column=1).value = erp_row['קוד פריט']
                    ws.cell(row=current_row_excel, column=2).value = erp_row['תיאור פריט']
                    ws.cell(row=current_row_excel, column=3).value = f"{score}%"
                    
                    # הכנת נתונים ללשונית הניתוח
                    alternatives = ", ".join([f"{m[0]} ({m[1]}%)" for m in matches[1:]])
                else:
                    ws.cell(row=current_row_excel, column=1).value = "לא נמצא"
                    ws.cell(row=current_row_excel, column=3).value = "0%"
                    alternatives = "אין הצעות קרובות"

                results_for_analysis.append({
                    'שורה באקסל': current_row_excel,
                    'תיאור מבוקש': query,
                    'נבחר ב-ERP': ws.cell(row=current_row_excel, column=2).value,
                    'ציון': score if matches else 0,
                    'הצעות חלופיות (3 הכי קרובות)': alternatives
                })
                current_row_excel += 1

            # יצירת לשונית ניתוח חדשה בקובץ
            if "ניתוח נתונים" in wb.sheetnames: del wb["ניתוח נתונים"]
            ws_analysis = wb.create_sheet("ניתוח נתונים")
            analysis_df = pd.DataFrame(results_for_analysis)
            for r in dataframe_to_rows(analysis_df, index=False, header=True):
                ws_analysis.append(r)

            # שמירה
            output = BytesIO()
            wb.save(output)
            
            st.success("העיבוד הושלם! הפורמט המקורי נשמר והתווספו הצעות חלופיות.")
            st.download_button("📥 הורד קובץ סופי מעובד", output.getvalue(), "Purchase_Order_Updated.xlsx")
